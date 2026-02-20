import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
import datetime
import sys
import csv
from pathlib import Path
import subprocess
import platform
import tempfile

# --- SCHEDA CONVERTITORE PDF (CODICE ESISTENTE) ---
class ConverterPanel(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        if getattr(sys, 'frozen', False):
            base_path = Path(sys.executable).parent
        else:
            base_path = Path(__file__).parent.resolve()
        
        self.work_dir = str(base_path.parent / "NP_STAMPE")
        try:
            if not os.path.exists(self.work_dir):
                os.makedirs(self.work_dir, exist_ok=True)
        except:
            self.work_dir = os.getcwd()

        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill="both", expand=True)

        dir_frame = ttk.LabelFrame(main_frame, text="Cartella di lavoro HTML", padding="5")
        dir_frame.pack(fill="x", pady=5)
        self.dir_label = ttk.Label(dir_frame, text=self.work_dir, relief="sunken", anchor="w")
        self.dir_label.pack(side="left", fill="x", expand=True, padx=(0, 5))
        ttk.Button(dir_frame, text="Cambia...", command=self.select_directory).pack(side="right")

        files_frame = ttk.LabelFrame(main_frame, text="File HTML disponibili per conversione", padding="5")
        files_frame.pack(fill="both", expand=True, pady=5)
        scrollbar = ttk.Scrollbar(files_frame)
        scrollbar.pack(side="right", fill="y")
        self.files_listbox = tk.Listbox(files_frame, selectmode="multiple", yscrollcommand=scrollbar.set, height=10)
        self.files_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.files_listbox.yview)

        btn_f = ttk.Frame(main_frame)
        btn_f.pack(pady=5)
        ttk.Button(btn_f, text="Seleziona tutti", command=lambda: self.files_listbox.selection_set(0, "end")).pack(side="left", padx=2)
        ttk.Button(btn_f, text="Deseleziona", command=lambda: self.files_listbox.selection_clear(0, "end")).pack(side="left", padx=2)
        ttk.Button(btn_f, text="Aggiorna lista", command=self.refresh_file_list).pack(side="left", padx=2)

        opt_f = ttk.LabelFrame(main_frame, text="Opzioni PDF", padding="5")
        opt_f.pack(fill="x", pady=5)
        self.mode = tk.StringVar(value="single")
        ttk.Radiobutton(opt_f, text="PDF separati (A4)", variable=self.mode, value="single").pack(anchor="w")
        ttk.Radiobutton(opt_f, text="Unisci in unico PDF", variable=self.mode, value="merged").pack(anchor="w")
        
        merge_n = ttk.Frame(opt_f)
        merge_n.pack(fill="x")
        ttk.Label(merge_n, text="Nome file unito:").pack(side="left", padx=5)
        self.merged_name = ttk.Entry(merge_n)
        self.merged_name.insert(0, "Progetto_Completo.pdf")
        self.merged_name.pack(side="left", fill="x", expand=True)

        self.btn_conv = ttk.Button(main_frame, text="AVVIA CONVERSIONE PDF", command=self.convert_to_pdf)
        self.btn_conv.pack(pady=10, ipady=5)

        self.status = ttk.Label(main_frame, text="Pronto", relief="sunken", anchor="w")
        self.status.pack(fill="x")
        self.prog = ttk.Progressbar(main_frame, mode='determinate')
        self.prog.pack(fill="x", pady=5)
        
        self.refresh_file_list()

    def select_directory(self):
        d = filedialog.askdirectory(initialdir=self.work_dir)
        if d:
            self.work_dir = d
            self.dir_label.config(text=d)
            self.refresh_file_list()

    def refresh_file_list(self):
        self.files_listbox.delete(0, "end")
        if os.path.exists(self.work_dir):
            files = sorted([f for f in os.listdir(self.work_dir) if f.lower().endswith('.html')])
            for f in files: self.files_listbox.insert("end", f)

    def get_chrome_path(self):
        sys_p = platform.system()
        if sys_p == "Darwin":
            ps = ["/Applications/Google Chrome.app/Contents/MacOS/Google Chrome", "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge"]
        elif sys_p == "Windows":
            ps = ["C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe", "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe", os.path.expanduser("~\\AppData\\Local\\Microsoft\\Edge\\Application\\msedge.exe")]
        else:
            ps = ["google-chrome", "microsoft-edge", "chromium"]
        for p in ps:
            if os.path.exists(p) if os.path.isabs(p) else True: return p
        return None

    def convert_to_pdf(self):
        sel = self.files_listbox.curselection()
        if not sel:
            messagebox.showwarning("Attenzione", "Seleziona almeno un file HTML")
            return
        
        c_path = self.get_chrome_path()
        if not c_path:
            messagebox.showerror("Errore", "Chrome o Edge non trovati. Installa un browser basato su Chromium.")
            return

        files = [self.files_listbox.get(i) for i in sel]
        mode = self.mode.get()
        self.prog['maximum'] = len(files)
        temp_pdfs = []

        try:
            for i, f in enumerate(files):
                h_p = os.path.abspath(os.path.join(self.work_dir, f))
                if mode == "single":
                    pdf_p = h_p.replace('.html', '.pdf')
                else:
                    t_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                    pdf_p = t_pdf.name
                    t_pdf.close()
                    temp_pdfs.append(pdf_p)

                self.status.config(text=f"Conversione {i+1}/{len(files)}...")
                self.update_idletasks()
                
                f_url = f"file:///{h_p.replace(os.sep, '/')}" if platform.system() == "Windows" else f"file://{h_p}"
                cmd = [c_path, "--headless", "--disable-gpu", "--no-pdf-header-footer", "--print-to-pdf=" + pdf_p, f_url]
                subprocess.run(cmd, check=True, timeout=30)
                self.prog['value'] = i + 1

            if mode == "merged":
                try:
                    from pypdf import PdfWriter, PdfReader
                except ImportError:
                     messagebox.showerror("Errore", "Libreria 'pypdf' mancante.\nInstalla con: pip install pypdf")
                     return

                out_n = self.merged_name.get()
                if not out_n.endswith('.pdf'): out_n += ".pdf"
                out_p = os.path.join(self.work_dir, out_n)
                writer = PdfWriter()
                for p in temp_pdfs:
                    r = PdfReader(p)
                    for page in r.pages: writer.add_page(page)
                with open(out_p, "wb") as o: writer.write(o)
                for p in temp_pdfs: os.unlink(p)
                messagebox.showinfo("Successo", f"PDF Unito creato: {out_n}")
            else:
                messagebox.showinfo("Successo", "Conversione completata.")
        except Exception as e:
            messagebox.showerror("Errore", str(e))
        finally:
            self.status.config(text="Pronto")
            self.prog['value'] = 0


# --- FUNZIONI DI UTILITÀ DEL PROGRAMMA PRINCIPALE ---
def format_currency(value):
    if value is None: value = 0.0
    return f"€ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def safe_float_convert(value_str):
    if not value_str: return 0.0
    try:
        clean = str(value_str).replace('€', '').strip()
        if '.' in clean and ',' in clean:
            clean = clean.replace('.', '').replace(',', '.')
        elif ',' in clean:
            clean = clean.replace(',', '.')
        return float(clean)
    except ValueError:
        return 0.0

# --- GESTIONE DATABASE ---
class Database:
    def __init__(self, db_name="np_zero.db"):
        self.db_name = db_name
        self.conn = sqlite3.connect(db_name)
        self.cursor = self.conn.cursor()
        self.create_tables()
        self.migrate_categories() # MIGRAZIONE AUTOMATICA ALL'AVVIO
        self.populate_defaults()

    def create_tables(self):
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS unita_misura (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codice TEXT UNIQUE,
                nome TEXT,
                descrizione TEXT
            )
        """)
        
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS progetti (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codice TEXT,
                titolo TEXT,
                cup TEXT,
                committente TEXT
            )
        """)

        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS nuovi_prezzi (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                progetto_id INTEGER,
                codice TEXT, 
                descrizione TEXT,
                unita_misura TEXT, 
                perc_spese_generali REAL DEFAULT 17.0,
                perc_sicurezza REAL DEFAULT 5.0,
                perc_utili REAL DEFAULT 10.0,
                prezzo_finale REAL DEFAULT 0.0,
                FOREIGN KEY(progetto_id) REFERENCES progetti(id) ON DELETE CASCADE
            )
        """)

        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS voci_costo (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                np_id INTEGER,
                ordine INTEGER,
                categoria TEXT, 
                descrizione TEXT,
                um TEXT,
                quantita REAL,
                prezzo_unitario REAL,
                FOREIGN KEY(np_id) REFERENCES nuovi_prezzi(id) ON DELETE CASCADE
            )
        """)
        self.conn.commit()
    
    # --- NUOVA FUNZIONE DI MIGRAZIONE ---
    def migrate_categories(self):
        """Converte le vecchie categorie nel nuovo standard Regionale Campania"""
        try:
            # Materiali -> Prodotti
            self.cursor.execute("UPDATE voci_costo SET categoria='Prodotti' WHERE categoria='Materiali'")
            # Mezzi e Noli -> Attrezzature
            self.cursor.execute("UPDATE voci_costo SET categoria='Attrezzature' WHERE categoria='Mezzi e Noli'")
            self.conn.commit()
        except Exception as e:
            print(f"Errore migrazione categorie: {e}")

    def populate_defaults(self):
        self.cursor.execute("SELECT count(*) FROM unita_misura")
        if self.cursor.fetchone()[0] == 0:
            defaults = [
                    ("cad.", "Cadauno", "Sanitari, infissi..."),
                    ("corpo", "A corpo", "Opere indivisibili"),
                    ("h", "Ora", "Manodopera..."),
                    ("kg", "Chilogrammo", "Acciaio..."),
                    ("m", "Metro lineare", "Tubazioni, cavi..."),
                    ("mc", "Metro cubo", "Getti in cls, scavi..."),
                    ("mq", "Metro quadrato", "Pavimenti, intonaci..."),
            ]
            self.cursor.executemany("INSERT INTO unita_misura (codice, nome, descrizione) VALUES (?, ?, ?)", defaults)
            self.conn.commit()

    def select_specific(self, table, columns, where_clause="", params=()):
        cols_str = ", ".join(columns)
        query = f"SELECT id, {cols_str} FROM {table} {where_clause}"
        self.cursor.execute(query, params)
        return self.cursor.fetchall()

    def fetch_all(self, table, where_clause="", params=()):
        query = f"SELECT * FROM {table} {where_clause}"
        self.cursor.execute(query, params)
        return self.cursor.fetchall()
    
    def fetch_one(self, table, where_clause="", params=()):
        query = f"SELECT * FROM {table} {where_clause}"
        self.cursor.execute(query, params)
        return self.cursor.fetchone()

    def get_max_order(self, np_id):
        query = "SELECT MAX(ordine) FROM voci_costo WHERE np_id=?"
        self.cursor.execute(query, (np_id,))
        val = self.cursor.fetchone()[0]
        return val if val is not None else 0

    def insert(self, table, data):
        columns = ', '.join(data.keys())
        placeholders = ', '.join(['?'] * len(data))
        values = list(data.values())
        query = f"INSERT INTO {table} ({columns}) VALUES ({placeholders})"
        self.cursor.execute(query, values)
        self.conn.commit()
        return self.cursor.lastrowid 

    def delete(self, table, record_id):
        self.cursor.execute(f"DELETE FROM {table} WHERE id=?", (record_id,))
        self.conn.commit()

    def update(self, table, record_id, data):
        set_clause = ', '.join([f"{k}=?" for k in data.keys()])
        values = list(data.values()) + [record_id]
        query = f"UPDATE {table} SET {set_clause} WHERE id=?"
        self.cursor.execute(query, values)
        self.conn.commit()

# --- FINESTRA DI IMPORTAZIONE ---
class ImportDialog(tk.Toplevel):
    def __init__(self, parent, db, current_project_id, on_import_callback):
        super().__init__(parent)
        self.title("Importa NP da altro progetto")
        self.geometry("600x400")
        self.db = db
        self.current_project_id = current_project_id
        self.on_import_callback = on_import_callback
        
        ttk.Label(self, text="1. Seleziona Progetto di Origine:", font=("Arial", 10, "bold")).pack(pady=5)
        
        self.combo_proj = ttk.Combobox(self, state="readonly", width=50)
        self.combo_proj.pack(pady=5)
        self.combo_proj.bind("<<ComboboxSelected>>", self.load_nps)
        
        ttk.Label(self, text="2. Seleziona NP da Copiare:", font=("Arial", 10, "bold")).pack(pady=5)
        
        cols = ("codice", "desc")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=10)
        self.tree.heading("codice", text="Codice")
        self.tree.heading("desc", text="Descrizione")
        self.tree.column("codice", width=100)
        self.tree.column("desc", width=400)
        self.tree.pack(fill="both", expand=True, padx=10, pady=5)
        
        btn_import = ttk.Button(self, text="IMPORTA SELEZIONATO", command=self.do_import)
        btn_import.pack(pady=10, fill="x", padx=20)
        
        self.load_projects()

    def load_projects(self):
        projs = self.db.fetch_all("progetti")
        self.proj_map = {}
        values = []
        for p in projs:
            if p[0] != self.current_project_id:
                label = f"{p[1]} - {p[2]}"
                values.append(label)
                self.proj_map[label] = p[0]
        self.combo_proj['values'] = values

    def load_nps(self, event):
        for row in self.tree.get_children(): self.tree.delete(row)
        selected_label = self.combo_proj.get()
        if not selected_label: return
        
        pid = self.proj_map[selected_label]
        nps = self.db.fetch_all("nuovi_prezzi", "WHERE progetto_id=?", (pid,))
        for np in nps:
            self.tree.insert("", "end", iid=np[0], values=(np[2], np[3]))

    def do_import(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Attenzione", "Seleziona un NP dalla lista")
            return
        
        source_np_id = sel[0]
        self.on_import_callback(source_np_id)
        self.destroy()

# --- CLASSE PER FILTRO STORICO ---
class HistoryDialog(tk.Toplevel):
    def __init__(self, parent, db, project_id, callback):
        super().__init__(parent)
        self.title("Cerca voce da storico progetto")
        self.geometry("700x500")
        self.db = db
        self.project_id = project_id
        self.callback = callback

        ttk.Label(self, text="Seleziona una voce utilizzata in altri NP di questo progetto:", font=("Arial", 10, "bold")).pack(pady=10)
        
        # Search box
        search_frame = ttk.Frame(self)
        search_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(search_frame, text="Filtra:").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.filter_list)
        ttk.Entry(search_frame, textvariable=self.search_var).pack(side="left", fill="x", expand=True, padx=5)

        # Treeview
        cols = ("cat", "desc", "um", "pu")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        self.tree.heading("cat", text="Categoria")
        self.tree.heading("desc", text="Descrizione")
        self.tree.heading("um", text="UM")
        self.tree.heading("pu", text="Prezzo Unitario")
        
        self.tree.column("cat", width=100)
        self.tree.column("desc", width=350)
        self.tree.column("um", width=50)
        self.tree.column("pu", width=80, anchor="e")
        
        self.tree.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.tree.bind("<Double-1>", self.on_double_click)

        btn_sel = ttk.Button(self, text="Usa Selezionato", command=self.use_selected)
        btn_sel.pack(pady=10)

        self.all_items = []
        self.load_data()

    def load_data(self):
        # Query DISTINCT con TRIM per pulire gli spazi ed eliminare i duplicati "sporchi"
        query = """
            SELECT DISTINCT TRIM(v.categoria), TRIM(v.descrizione), TRIM(v.um), v.prezzo_unitario
            FROM voci_costo v
            JOIN nuovi_prezzi n ON v.np_id = n.id
            WHERE n.progetto_id = ?
            ORDER BY 1, 2
        """
        rows = self.db.cursor.execute(query, (self.project_id,)).fetchall()
        self.all_items = rows
        self.update_tree(rows)

    def filter_list(self, *args):
        search_term = self.search_var.get().lower()
        filtered = [item for item in self.all_items if search_term in item[1].lower() or search_term in item[0].lower()]
        self.update_tree(filtered)

    def update_tree(self, items):
        for row in self.tree.get_children(): self.tree.delete(row)
        for item in items:
            pu_fmt = f"€ {item[3]:.3f}".replace(".", ",")
            self.tree.insert("", "end", values=(item[0], item[1], item[2], pu_fmt), tags=(item,))

    def use_selected(self):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0])['values']
        cat = vals[0]
        desc = vals[1]
        um = vals[2]
        
        pu_str = vals[3].replace("€", "").strip().replace(".", "").replace(",", ".")
        pu = float(pu_str)

        self.callback(cat, desc, um, pu)
        self.destroy()

    def on_double_click(self, event):
        self.use_selected()

# --- PANNELLO CRUD (GESTIONE LAYOUT MULTIPLI) ---
class CrudPanel(ttk.Frame):
    def __init__(self, parent, db, table_name, fields, callbacks=None):
        super().__init__(parent)
        self.db = db
        self.table_name = table_name
        self.fields = fields 
        self.callbacks = callbacks or {}
        self.entries = {}
        self.buttons = {}

        self.frame_top = ttk.LabelFrame(self, text="Dati Inserimento", padding=10)
        self.frame_top.pack(side="top", fill="x", padx=10, pady=5)
        
        if self.table_name == "nuovi_prezzi":
            self.setup_custom_np_ui()
        elif self.table_name == "progetti":
            self.setup_custom_progetti_ui()
        else:
            self.setup_generic_ui()

        self.frame_mid = ttk.Frame(self)
        self.frame_mid.pack(side="top", fill="both", expand=True, padx=10, pady=5)

        cols = [f[0] for f in self.fields]
        self.tree = ttk.Treeview(self.frame_mid, columns=cols, show="headings")
        
        for field, label, width in self.fields:
            align = "w"
            if field == "prezzo_finale": align = "e"
            self.tree.heading(field, text=label, 
                              command=lambda c=field: self.sort_column(c, False))
            self.tree.column(field, width=width, anchor=align)
        
        self.tree.pack(side="left", fill="both", expand=True)
        
        sb = ttk.Scrollbar(self.frame_mid, orient="vertical", command=self.tree.yview)
        sb.pack(side="left", fill="y")
        self.tree.configure(yscroll=sb.set)

        self.frame_btns = ttk.Frame(self.frame_mid)
        self.frame_btns.pack(side="right", fill="y", padx=5)

        self.buttons['add'] = ttk.Button(self.frame_btns, text="Nuovo", command=self.add_record)
        self.buttons['add'].pack(fill="x", pady=2)
        self.buttons['upd'] = ttk.Button(self.frame_btns, text="Modifica", command=self.update_record)
        self.buttons['upd'].pack(fill="x", pady=2)
        self.buttons['del'] = ttk.Button(self.frame_btns, text="Cancella", command=self.delete_record)
        self.buttons['del'].pack(fill="x", pady=2)
        self.buttons['dup'] = ttk.Button(self.frame_btns, text="Duplica", command=self.duplicate_record)
        self.buttons['dup'].pack(fill="x", pady=2)
        self.buttons['imp'] = ttk.Button(self.frame_btns, text="Importa da...", command=lambda: print("Implementare Wrapper")) 
        self.buttons['imp'].pack(fill="x", pady=2)
        
        ttk.Separator(self.frame_btns, orient="horizontal").pack(fill="x", pady=5)
        self.buttons['clr'] = ttk.Button(self.frame_btns, text="Pulisci", command=self.clear_fields)
        self.buttons['clr'].pack(fill="x", pady=10)

        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Double-1>", self.on_double_click)
        
        self.refresh_data()

    def setup_generic_ui(self):
        for i, (field_name, label_text, width) in enumerate(self.fields):
            if field_name == "prezzo_finale": continue
            container = ttk.Frame(self.frame_top)
            container.pack(side="left", padx=5, fill="y")
            ttk.Label(container, text=label_text, font=("Arial", 9, "bold")).pack(side="top", anchor="w")
            entry = ttk.Entry(container, width=width // 8) 
            entry.pack(side="top", fill="x")
            self.entries[field_name] = entry

    def setup_custom_progetti_ui(self):
        row1 = ttk.Frame(self.frame_top)
        row1.pack(side="top", fill="x", pady=2)

        def add_field(parent, key, label, width_factor):
            f = ttk.Frame(parent)
            f.pack(side="left", padx=(0, 15), fill="x", expand=False)
            ttk.Label(f, text=label, font=("Arial", 9, "bold")).pack(anchor="w")
            e = ttk.Entry(f, width=width_factor)
            e.pack(fill="x")
            self.entries[key] = e

        add_field(row1, 'codice', 'Codice', 15)
        add_field(row1, 'cup', 'CUP', 15)
        
        f_comm = ttk.Frame(row1)
        f_comm.pack(side="left", padx=0, fill="x", expand=True)
        ttk.Label(f_comm, text="Committente", font=("Arial", 9, "bold")).pack(anchor="w")
        self.entries['committente'] = ttk.Entry(f_comm)
        self.entries['committente'].pack(fill="x")

        row2 = ttk.Frame(self.frame_top)
        row2.pack(side="top", fill="x", pady=5)
        
        ttk.Label(row2, text="Titolo Intervento", font=("Arial", 9, "bold")).pack(anchor="w")
        self.entries['titolo'] = ttk.Entry(row2)
        self.entries['titolo'].pack(fill="x")

    def setup_custom_np_ui(self):
        row1 = ttk.Frame(self.frame_top)
        row1.pack(side="top", fill="x", pady=2)
        
        f_cod = ttk.Frame(row1)
        f_cod.pack(side="left", padx=(0, 20))
        ttk.Label(f_cod, text="Codice NP", font=("Arial", 9, "bold")).pack(anchor="w")
        self.entries['codice'] = ttk.Entry(f_cod, width=15)
        self.entries['codice'].pack(fill="x")

        f_um = ttk.Frame(row1)
        f_um.pack(side="left")
        ttk.Label(f_um, text="Unità di Misura", font=("Arial", 9, "bold")).pack(anchor="w")
        self.entries['unita_misura'] = ttk.Combobox(f_um, width=15, postcommand=self.populate_um_combo)
        self.entries['unita_misura'].pack(fill="x")

        row2 = ttk.Frame(self.frame_top)
        row2.pack(side="top", fill="x", pady=5)
        
        ttk.Label(row2, text="Descrizione NP", font=("Arial", 9, "bold")).pack(anchor="w")
        
        txt_frame = ttk.Frame(row2)
        txt_frame.pack(fill="x", expand=True)
        
        self.entries['descrizione'] = tk.Text(txt_frame, height=4, font=("Arial", 9))
        self.entries['descrizione'].pack(side="left", fill="x", expand=True)
        
        sb = ttk.Scrollbar(txt_frame, orient="vertical", command=self.entries['descrizione'].yview)
        sb.pack(side="right", fill="y")
        self.entries['descrizione'].config(yscrollcommand=sb.set)

    def populate_um_combo(self):
        if 'unita_misura' in self.entries and isinstance(self.entries['unita_misura'], ttk.Combobox):
            ums = self.db.fetch_all("unita_misura")
            values = [u[1] for u in ums]
            self.entries['unita_misura']['values'] = values

    def sort_column(self, col, reverse):
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        def sort_key(t):
            val = t[0]
            if not val: return ""
            try:
                clean = str(val).replace('€', '').strip()
                if '.' in clean and ',' in clean: clean = clean.replace('.', '').replace(',', '.')
                elif ',' in clean: clean = clean.replace(',', '.')
                return float(clean)
            except ValueError:
                return str(val).lower()
        l.sort(key=sort_key, reverse=reverse)
        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)
        self.tree.heading(col, command=lambda: self.sort_column(col, not reverse))

    def get_data_from_ui(self):
        data = {}
        for field, _, _ in self.fields:
            if field in self.entries:
                widget = self.entries[field]
                if isinstance(widget, tk.Text):
                    val = widget.get("1.0", "end-1c").strip()
                else:
                    val = widget.get().strip()
                data[field] = val
        return data

    def clear_fields(self):
        for widget in self.entries.values():
            if isinstance(widget, tk.Text):
                widget.delete("1.0", tk.END)
            else:
                widget.delete(0, tk.END)

    def refresh_data(self, condition="", params=()):
        for row in self.tree.get_children(): self.tree.delete(row)
        target_columns = [f[0] for f in self.fields]
        rows = self.db.select_specific(self.table_name, target_columns, condition, params)
        for r in rows:
            vals = list(r[1:])
            for idx, f in enumerate(self.fields):
                if f[0] == "prezzo_finale" and idx < len(vals):
                     vals[idx] = format_currency(vals[idx] or 0.0)
            self.tree.insert("", "end", iid=r[0], values=vals)

    def add_record(self):
        try:
            data = self.get_data_from_ui()
            if not any(data.values()): return
            self.db.insert(self.table_name, data)
            self.refresh_data() 
            self.clear_fields()
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def update_record(self):
        selected = self.tree.selection()
        if not selected: return
        try:
            data = self.get_data_from_ui()
            self.db.update(self.table_name, selected[0], data)
            self.refresh_data() 
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def delete_record(self):
        selected = self.tree.selection()
        if not selected: return
        if messagebox.askyesno("Conferma", "Cancellare record?"):
            self.db.delete(self.table_name, selected[0])
            self.refresh_data() 
            self.clear_fields()

    def duplicate_record(self):
        selected = self.tree.selection()
        if not selected: return
        data = self.get_data_from_ui()
        if 'codice' in data: data['codice'] += "_cp"
        if 'titolo' in data: data['titolo'] += " (Copia)"
        try:
            self.db.insert(self.table_name, data)
            self.refresh_data() 
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def on_select(self, event):
        selected = self.tree.selection()
        if not selected: return
        item = self.tree.item(selected[0])
        values = item['values']
        self.clear_fields()
        
        col_idx = 0
        for field, _, _ in self.fields:
            if field in self.entries and col_idx < len(values):
                widget = self.entries[field]
                val = values[col_idx]
                
                if isinstance(widget, tk.Text):
                    widget.insert("1.0", str(val))
                elif isinstance(widget, ttk.Combobox):
                     widget.set(str(val))
                else:
                    widget.insert(0, str(val))
            col_idx += 1
            
        if 'on_select' in self.callbacks:
            self.callbacks['on_select'](selected[0])
            
    def on_double_click(self, event):
        selected = self.tree.selection()
        if not selected: return
        if 'on_double_click' in self.callbacks:
            self.callbacks['on_double_click'](selected[0])

# --- SCHEDA STAMPE (HTML) ---
class PrintPanel(ttk.Frame):
    def __init__(self, parent, db):
        super().__init__(parent)
        self.db = db
        self.current_np_id = None
        self.print_proj_map = {}

        self.lbl_title = ttk.Label(self, text="Stampa Singola NP (HTML)", font=("Arial", 14, "bold"))
        self.lbl_title.pack(pady=(20, 5))
        
        self.info_lbl = ttk.Label(self, text="Seleziona un NP nella Scheda 2 per abilitare la stampa singola.", font=("Arial", 10))
        self.info_lbl.pack(pady=5)
        
        self.btn_print = ttk.Button(self, text="GENERA STAMPA NP SELEZIONATO", command=self.generate_html, width=40)
        self.btn_print.pack(pady=5, ipady=5)

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=50, pady=20)

        self.lbl_batch = ttk.Label(self, text="Esportazione Massiva per Intervento", font=("Arial", 14, "bold"))
        self.lbl_batch.pack(pady=5)
        
        ttk.Label(self, text="Seleziona un progetto per generare stampe o esportare in Excel:").pack(pady=5)
        
        self.combo_stampa_progetti = ttk.Combobox(self, state="readonly", width=60, postcommand=self.load_projects_for_print)
        self.combo_stampa_progetti.pack(pady=5)
        
        # Frame per affiancare i due pulsanti massivi (HTML e EXCEL)
        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=10)
        
        self.btn_print_all = ttk.Button(btn_frame, text="GENERA TUTTI GLI HTML", command=self.generate_batch_html, width=30)
        self.btn_print_all.pack(side="left", padx=10, ipady=5)

        self.btn_export_excel = ttk.Button(btn_frame, text="ESPORTA PROGETTO IN EXCEL", command=self.generate_excel_export, width=30)
        self.btn_export_excel.pack(side="left", padx=10, ipady=5)

        self.lbl_status = ttk.Label(self, text="", foreground="green", font=("Arial", 9, "italic"))
        self.lbl_status.pack(pady=20)

    def set_current_np(self, np_id):
        self.current_np_id = np_id
        if np_id:
            np = self.db.fetch_one("nuovi_prezzi", "WHERE id=?", (np_id,))
            self.info_lbl.config(text=f"NP Selezionato: {np[2]} - {np[3]}")
        else:
            self.info_lbl.config(text="Nessun NP selezionato nella Scheda 2.")

    def load_projects_for_print(self):
        projs = self.db.fetch_all("progetti")
        values = []
        self.print_proj_map = {}
        for p in projs:
            label = f"{p[1]} - {p[2]}"
            values.append(label)
            self.print_proj_map[label] = p[0]
        self.combo_stampa_progetti['values'] = values

    def generate_batch_html(self):
        selected_label = self.combo_stampa_progetti.get()
        if not selected_label:
            messagebox.showwarning("Attenzione", "Seleziona un progetto dal menu.")
            return
        
        proj_id = self.print_proj_map[selected_label]
        nps = self.db.fetch_all("nuovi_prezzi", "WHERE progetto_id=?", (proj_id,))
        
        if not nps:
            messagebox.showinfo("Info", "Nessun NP associato a questo progetto.")
            return

        if messagebox.askyesno("Conferma", f"Verranno generati {len(nps)} file HTML. Continuare?"):
            original_np_id = self.current_np_id
            count = 0
            for np in nps:
                self.current_np_id = np[0]
                self.generate_html()
                count += 1
            self.current_np_id = original_np_id
            self.lbl_status.config(text=f"Operazione completata: {count} stampe generate.")
            messagebox.showinfo("Successo", f"Generati {count} file nella cartella NP_STAMPE.")

    def generate_html(self):
        if not hasattr(self, 'current_np_id') or not self.current_np_id:
            messagebox.showerror("Errore", "Nessun NP selezionato.")
            return

        np = self.db.fetch_one("nuovi_prezzi", "WHERE id=?", (self.current_np_id,))
        proj = self.db.fetch_one("progetti", "WHERE id=?", (np[1],))
        items = self.db.fetch_all("voci_costo", "WHERE np_id=? ORDER BY categoria, ordine", (self.current_np_id,))

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_NP_{np[2]}.html"
        
        if getattr(sys, 'frozen', False):
            current_path = Path(sys.executable).parent
        else:
            current_path = Path(__file__).parent.resolve()
        
        target_dir_path = current_path.parent / "NP_STAMPE"
        target_dir_path.mkdir(parents=True, exist_ok=True)
        full_path = str(target_dir_path / filename)

        css = """
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; padding: 10px; color: #333; max-width: 900px; margin: auto; }
        h1 { color: #0055aa; border-bottom: 2px solid #0055aa; padding-bottom: 10px; font-size: 24px; }
        h2 { background-color: #f2f2f2; padding: 5px; margin-top: 30px; border-left: 5px solid #0055aa; font-size: 18px; }
        .header-box { background: #eef; padding: 10px; border-radius: 5px; margin-bottom: 20px; border: 1px solid #ccd; }
        table { width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 14px; }
        th, td { border: 1px solid #ddd; padding: 3px; text-align: left; }
        th { background-color: #0055aa; color: white; text-transform: uppercase; font-size: 12px; }
        .num { text-align: right; }
        .footer { font-size: 0.8em; color: #777; margin-top: 40px; text-align: center; border-top: 1px solid #ddd; padding-top: 10px; }
        .small-note { font-size: 0.85em; color: #666; margin-top: 2px; display: block; }
        """

        html = f"""
        <!DOCTYPE html>
        <html lang="it">
        <head>
            <meta charset="UTF-8">
            <title>Analisi Prezzo {np[2]}</title>
            <style>{css}</style>
        </head>
        <body>
            <h1>ANALISI NUOVO PREZZO: {np[2]}</h1>
            <div class="header-box">
                <table style="width: 100%; border-collapse: collapse; border: none;">
                    <tr>
                        <td colspan="2" style="border: none; padding: 2px 0;">
                            <strong>Descrizione:</strong> {np[3]}
                        </td>
                    </tr>
                    <tr><td colspan="2" style="border: none;"></td></tr>
                    <tr style="font-size: 1.1em;">
                        <td style="border: none; text-align: left; padding: 0; width: 50%;">
                            <strong>Unità di Misura: {np[4]}</strong>
                        </td>
                        <td style="border: none; text-align: right; padding: 0; width: 50%;">
                            <strong>Prezzo d'applicazione: {format_currency(np[8])}</strong>
                        </td>
                    </tr>
                </table>
            </div>
            <h2>Dettaglio Voci Costitutive</h2>
            <table>
                <thead>
                    <tr>
                        <th width="5%">N.</th>
                        <th>Descrizione</th>
                        <th width="5%">U.M.</th>
                        <th width="10%">Quantità</th>
                        <th width="15%">P. Unitario</th>
                        <th width="15%">Totale</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        current_cat = ""
        tot_a = 0.0
        tot_manodopera = 0.0

        for item in items:
            cat, desc, um, q, pu = item[3], item[4], item[5], item[6], item[7]
            tot = q * pu
            tot_a += tot
            
            if "Manodopera" in cat:
                tot_manodopera += tot

            if cat != current_cat:
                html += f"""<tr><td colspan="6" style="background-color:#e0e0e0; font-weight:bold; color:#333; padding-left: 10px;">{cat.upper()}</td></tr>"""
                current_cat = cat

            html += f"""
                <tr>
                    <td></td>
                    <td>{desc}</td>
                    <td>{um}</td>
                    <td class="num">{q:.3f}</td>
                    <td class="num">€ {pu:.3f}</td>
                    <td class="num">€ {tot:.2f}</td>
                </tr>
            """

        sg_val = tot_a * (np[5]/100)
        sic_val = sg_val * (np[6]/100)
        utili_val = (tot_a + sg_val) * (np[7]/100)
        perc_man = (tot_manodopera / np[8] * 100) if np[8] > 0 else 0
        perc_sic = (sic_val / np[8] * 100) if np[8] > 0 else 0

        html += f"""
                </tbody>
            </table>

            <h2>Riepilogo economico</h2>
            <table style="width: 100%; border: 2px solid #ddd;">
                <tr>
                    <td>(A) Sommano</td>
                    <td class="num">{format_currency(tot_a)}</td>
                </tr>
                <tr>
                    <td style="white-space: nowrap;">
                        (B) Spese Generali = {np[5]}% di A <span class="small-note" style="display: inline !important;"> - compresi oneri di sicurezza afferenti all'impresa per euro {format_currency(sic_val)} = {np[6]}% di B</span>
                    </td>
                    <td class="num">{format_currency(sg_val)}</td>
                </tr>
                <tr>
                    <td>(C) Utile d'Impresa = {np[7]}% di (A + B)</td>
                    <td class="num">{format_currency(utili_val)}</td>
                </tr>
                <tr style="background-color: #dbeeff;">
                    <td style="font-weight: bold;">(D) TOTALE (A + B + C)</td>
                    <td class="num" style="font-weight: bold;">{format_currency(np[8])}</td>
                </tr>
            </table>

            <h2>Riepilogo incidenze</h2>
            <table style="width: 100%; border: 1px solid #ddd;">
                <tr>
                    <td>Incidenza Manodopera sul totale</td>
                    <td class="num"> {perc_man:.2f}%</td>
                    <td class="num">{format_currency(tot_manodopera)}</td>
                </tr>
                <tr>
                    <td>Incidenza Oneri di Sicurezza sul totale</td>
                    <td class="num">{perc_sic:.2f}%</td>
                    <td class="num">{format_currency(sic_val)}</td>
                </tr>
            </table>

            <div class="footer">Generato il {datetime.datetime.now().strftime("%d/%m/%Y")}</div>
        </body>
        </html>
        """

        with open(full_path, "w", encoding="utf-8") as f:
            f.write(html)
        
        self.lbl_status.config(text=f"Stampa salvata in:\n{full_path}")
        try:
            if os.name == 'nt': os.startfile(full_path)
            else: os.system(f"open '{full_path}'")
        except:
            pass

# --- AGGIUNGI QUESTO NUOVO METODO ALLA FINE DELLA CLASSE PrintPanel ---
    def generate_excel_export(self):
        selected_label = self.combo_stampa_progetti.get()
        if not selected_label:
            messagebox.showwarning("Attenzione", "Seleziona un progetto dal menu.")
            return
        
        proj_id = self.print_proj_map[selected_label]
        proj = self.db.fetch_one("progetti", "WHERE id=?", (proj_id,))
        nps = self.db.fetch_all("nuovi_prezzi", "WHERE progetto_id=?", (proj_id,))
        
        if not nps:
            messagebox.showinfo("Info", "Nessun NP associato a questo progetto.")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            messagebox.showerror("Errore", "Libreria 'openpyxl' mancante.\nAssicurati di averla installata.")
            return

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        # Puliamo il nome del progetto da caratteri non validi per il nome file
        safe_proj_name = "".join([c for c in proj[2] if c.isalnum() or c in (' ', '_')]).rstrip()
        filename = f"{safe_proj_name}_{timestamp}.xlsx"
        
        if getattr(sys, 'frozen', False):
            current_path = Path(sys.executable).parent
        else:
            current_path = Path(__file__).parent.resolve()
        
        target_dir_path = current_path.parent / "NP_EXPORT"
        target_dir_path.mkdir(parents=True, exist_ok=True)
        full_path = str(target_dir_path / filename)

        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Rimuove il foglio di default vuoto

        # Definizione Stili Excel
        header_font = Font(bold=True, size=14, color="0055AA")
        bold_font = Font(bold=True)
        white_bold_font = Font(bold=True, color="FFFFFF")
        fill_blue = PatternFill(start_color="0055AA", end_color="0055AA", fill_type="solid")
        fill_light_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        fill_grey = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        for np in nps:
            # Crea un nome foglio valido per Excel (max 31 caratteri e senza simboli strani)
            sheet_name = str(np[2])[:31] 
            for char in ['*', ':', '?', '/', '\\', '[', ']']:
                sheet_name = sheet_name.replace(char, '')
            ws = wb.create_sheet(title=sheet_name)

            # Dimensioni colonne
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15

            # Intestazione NP
            ws.merge_cells('A1:F1')
            ws['A1'] = f"ANALISI NUOVO PREZZO: {np[2]}"
            ws['A1'].font = header_font
            
            ws.merge_cells('A3:F3')
            ws['A3'] = f"Descrizione: {np[3]}"
            ws['A3'].font = bold_font
            ws['A3'].alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[3].height = 40

            ws.merge_cells('A5:C5')
            ws['A5'] = f"Unità di Misura: {np[4]}"
            ws['A5'].font = bold_font

            ws.merge_cells('D5:F5')
            ws['D5'] = f"Prezzo d'applicazione: {format_currency(np[8])}"
            ws['D5'].font = bold_font
            ws['D5'].alignment = align_right

            # Intestazioni Tabella Voci di Costo
            row_num = 7
            headers = ["N.", "Descrizione", "U.M.", "Quantità", "P. Unitario", "Totale"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = white_bold_font
                cell.fill = fill_blue
                cell.alignment = align_center
                cell.border = border_thin

            row_num += 1
            
            items = self.db.fetch_all("voci_costo", "WHERE np_id=? ORDER BY categoria, ordine", (np[0],))
            current_cat = ""
            tot_a = 0.0
            tot_manodopera = 0.0

            for item in items:
                cat, desc, um, q, pu = item[3], item[4], item[5], item[6], item[7]
                tot = q * pu
                tot_a += tot
                
                if "Manodopera" in cat:
                    tot_manodopera += tot

                if cat != current_cat:
                    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
                    cell = ws.cell(row=row_num, column=1, value=cat.upper())
                    cell.font = bold_font
                    cell.fill = fill_grey
                    cell.border = border_thin
                    for c in range(2, 7):
                        ws.cell(row=row_num, column=c).border = border_thin
                    row_num += 1
                    current_cat = cat

                # Scrittura Riga Voce
                ws.cell(row=row_num, column=1, value="").border = border_thin
                ws.cell(row=row_num, column=2, value=desc).border = border_thin
                
                c_um = ws.cell(row=row_num, column=3, value=um)
                c_um.alignment = align_center
                c_um.border = border_thin
                
                c_q = ws.cell(row=row_num, column=4, value=q)
                c_q.number_format = '#,##0.000'
                c_q.border = border_thin
                
                c_pu = ws.cell(row=row_num, column=5, value=pu)
                c_pu.number_format = '€ #,##0.000'
                c_pu.border = border_thin
                
                c_tot = ws.cell(row=row_num, column=6, value=tot)
                c_tot.number_format = '€ #,##0.00'
                c_tot.border = border_thin
                
                row_num += 1

            # Riepilogo Economico
            row_num += 2
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            cell = ws.cell(row=row_num, column=1, value="Riepilogo economico")
            cell.font = bold_font
            cell.fill = fill_light_grey
            row_num += 1

            sg_val = tot_a * (np[5]/100)
            sic_val = sg_val * (np[6]/100)
            utili_val = (tot_a + sg_val) * (np[7]/100)
            
            perc_man = (tot_manodopera / np[8] * 100) if np[8] > 0 else 0
            perc_sic = (sic_val / np[8] * 100) if np[8] > 0 else 0

            def add_summary_row(r_idx, test_left, val_right, bold=False):
                ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=5)
                c_l = ws.cell(row=r_idx, column=1, value=test_left)
                c_r = ws.cell(row=r_idx, column=6, value=val_right)
                c_r.number_format = '€ #,##0.00'
                if bold:
                    c_l.font = bold_font
                    c_r.font = bold_font
                return r_idx + 1

            row_num = add_summary_row(row_num, "(A) Sommano", tot_a)
            row_num = add_summary_row(row_num, f"(B) Spese Generali = {np[5]}% di A (compresi oneri sicurezza € {sic_val:.2f} = {np[6]}% di B)", sg_val)
            row_num = add_summary_row(row_num, f"(C) Utile d'Impresa = {np[7]}% di (A + B)", utili_val)
            row_num = add_summary_row(row_num, "(D) TOTALE (A + B + C)", np[8], bold=True)

            row_num += 1
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            cell = ws.cell(row=row_num, column=1, value="Riepilogo incidenze")
            cell.font = bold_font
            cell.fill = fill_light_grey
            row_num += 1

            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=1, value="Incidenza Manodopera sul totale")
            ws.cell(row=row_num, column=5, value=f"{perc_man:.2f}%").alignment = align_right
            c_m = ws.cell(row=row_num, column=6, value=tot_manodopera)
            c_m.number_format = '€ #,##0.00'
            row_num += 1

            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=1, value="Incidenza Oneri di Sicurezza sul totale")
            ws.cell(row=row_num, column=5, value=f"{perc_sic:.2f}%").alignment = align_right
            c_s = ws.cell(row=row_num, column=6, value=sic_val)
            c_s.number_format = '€ #,##0.00'

        wb.save(full_path)
        self.lbl_status.config(text=f"Export Excel completato:\n{full_path}")
        messagebox.showinfo("Successo", f"File Excel generato con successo nella cartella:\nNP_EXPORT")
        
        # Apri la cartella al termine dell'operazione
        try:
            if os.name == 'nt': os.startfile(target_dir_path)
            else: os.system(f"open '{target_dir_path}'")
        except:
            pass

# --- APP PRINCIPALE ---
class NPApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gestione Nuovi Prezzi (NP)")
        self.geometry("1100x800")
        
        self.db = Database()
        
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_change)

        # TAB 1: PROGETTI
        self.tab_progetti = CrudPanel(
            self.notebook, self.db, "progetti", 
            [("codice", "Codice", 100), ("titolo", "Titolo Intervento", 300), ("cup", "CUP", 100), ("committente", "Committente", 200)],
            callbacks={'on_select': self.on_project_select, 'on_double_click': self.goto_tab_np}
        )
        self.notebook.add(self.tab_progetti, text="1. Progetti")

        # TAB 2: ELENCO NP
        self.frame_tab2 = ttk.Frame(self.notebook) 
        self.notebook.add(self.frame_tab2, text="2. Elenco NP")
        
        self.lbl_project_title = ttk.Label(self.frame_tab2, text="Nessun Progetto Selezionato", font=("Arial", 12, "bold"), foreground="#444", padding=10)
        self.lbl_project_title.pack(side="top", fill="x")

        self.current_project_id = None
        fields_np = [
            ("codice", "Codice NP", 100), 
            ("descrizione", "Descrizione NP", 400), 
            ("unita_misura", "Unità di Misura", 100),
            ("prezzo_finale", "Prezzo Finale", 120)
        ]
        
        self.tab_np = CrudPanel(
            self.frame_tab2, self.db, "nuovi_prezzi",
            fields_np,
            callbacks={'on_select': self.on_np_select, 'on_double_click': self.goto_tab_details}
        )
        self.tab_np.pack(fill="both", expand=True)
        
        def add_np_wrapper():
            if not self.current_project_id:
                messagebox.showwarning("Attenzione", "Seleziona prima un Progetto dalla Scheda 1")
                return
            data = self.tab_np.get_data_from_ui()
            data['progetto_id'] = self.current_project_id
            data['perc_spese_generali'] = 17.0
            data['perc_sicurezza'] = 5.0
            data['perc_utili'] = 10.0
            data['prezzo_finale'] = 0.0
            try:
                self.db.insert("nuovi_prezzi", data)
                self.tab_np.refresh_data("WHERE progetto_id=?", (self.current_project_id,))
                self.tab_np.clear_fields()
            except Exception as e:
                messagebox.showerror("Errore", str(e))
        
        def duplicate_np_deep_wrapper():
            self._copy_np_logic(is_import=False)

        def import_np_wrapper():
            if not self.current_project_id:
                messagebox.showwarning("Attenzione", "Seleziona prima un Progetto di destinazione")
                return
            ImportDialog(self, self.db, self.current_project_id, self._import_np_callback)

        def update_np_wrapper():
            self.tab_np.__class__.update_record(self.tab_np)
            if self.current_project_id: self.tab_np.refresh_data("WHERE progetto_id=?", (self.current_project_id,))

        def delete_np_wrapper():
            self.tab_np.__class__.delete_record(self.tab_np)
            if self.current_project_id: self.tab_np.refresh_data("WHERE progetto_id=?", (self.current_project_id,))

        self.tab_np.buttons['add'].config(command=add_np_wrapper)
        self.tab_np.buttons['upd'].config(command=update_np_wrapper)
        self.tab_np.buttons['del'].config(command=delete_np_wrapper)
        self.tab_np.buttons['dup'].config(command=duplicate_np_deep_wrapper)
        self.tab_np.buttons['imp'].config(command=import_np_wrapper)

        # TAB 3: MODIFICA NP
        self.frame_det_np = ttk.Frame(self.notebook)
        self.notebook.add(self.frame_det_np, text="3. Modifica NP")
        self.setup_details_tab()

        # TAB 4: AMMINISTRAZIONE
        self.tab_admin = CrudPanel(
            self.notebook, self.db, "unita_misura",
            [("codice", "Codice", 80), ("nome", "Unità", 150), ("descrizione", "Utilizzo", 400)]
        )
        self.notebook.add(self.tab_admin, text="4. Unità di Misura")
        
        # --- FUNZIONI EXPORT/IMPORT CSV (TAB 4) ---
        def export_um_csv():
            path = filedialog.asksaveasfilename(
                title="Esporta Unità di Misura",
                defaultextension=".csv",
                filetypes=[("File CSV", "*.csv")]
            )
            if not path: return
            try:
                rows = self.db.fetch_all("unita_misura")
                with open(path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f, delimiter=';')
                    writer.writerow(["CODICE", "NOME", "DESCRIZIONE"])
                    for r in rows:
                        writer.writerow([r[1], r[2], r[3]])
                messagebox.showinfo("Export", "Esportazione CSV completata con successo!")
            except Exception as e:
                messagebox.showerror("Errore Export", str(e))

        def import_um_csv():
            path = filedialog.askopenfilename(
                title="Importa Unità di Misura",
                filetypes=[("File CSV", "*.csv")]
            )
            if not path: return
            try:
                count = 0
                with open(path, newline='', encoding='utf-8') as f:
                    sample = f.read(1024)
                    f.seek(0)
                    sniffer = csv.Sniffer()
                    try:
                        dialect = sniffer.sniff(sample, delimiters=";,")
                    except:
                        dialect = 'excel' 

                    reader = csv.reader(f, dialect)
                    try:
                        header = next(reader)
                    except StopIteration:
                        return 

                    for row in reader:
                        if len(row) < 2: continue 
                        cod = row[0].strip()
                        nom = row[1].strip()
                        des = row[2].strip() if len(row) > 2 else ""
                        existing = self.db.fetch_one("unita_misura", "WHERE codice=?", (cod,))
                        if not existing:
                            self.db.insert("unita_misura", {'codice': cod, 'nome': nom, 'descrizione': des})
                            count += 1
                
                self.tab_admin.refresh_data()
                messagebox.showinfo("Import", f"Importazione completata.\nInseriti {count} nuovi record.")
            except Exception as e:
                messagebox.showerror("Errore Import", f"Errore lettura CSV:\n{str(e)}\n\nAssicurati che il formato sia: Codice;Nome;Descrizione")

        ttk.Separator(self.tab_admin.frame_btns, orient="horizontal").pack(fill="x", pady=5)
        btn_exp = ttk.Button(self.tab_admin.frame_btns, text="Esporta CSV", command=export_um_csv)
        btn_exp.pack(fill="x", pady=2)
        btn_imp = ttk.Button(self.tab_admin.frame_btns, text="Importa CSV", command=import_um_csv)
        btn_imp.pack(fill="x", pady=2)

        # TAB 5: STAMPE
        self.tab_print = PrintPanel(self.notebook, self.db)
        self.notebook.add(self.tab_print, text="5. Stampe")

        # --- TAB 6: CONVERTITORE PDF ---
        self.tab_converter = ConverterPanel(self.notebook)
        self.notebook.add(self.tab_converter, text="6. Convertitore PDF")

    # --- LOGICA DI COPIA ---
    def _import_np_callback(self, source_np_id):
        self._copy_np_logic(is_import=True, source_id_override=source_np_id)

    def _copy_np_logic(self, is_import=False, source_id_override=None):
        if not self.current_project_id: return
        
        if is_import:
            source_id = source_id_override
        else:
            selected = self.tab_np.tree.selection()
            if not selected: return
            source_id = selected[0]
            
        src_rec = self.db.fetch_one("nuovi_prezzi", "WHERE id=?", (source_id,))
        
        data = {
            'progetto_id': self.current_project_id,
            'codice': src_rec[2] + ("_imp" if is_import else "_cp"),
            'descrizione': src_rec[3] + (" (Import)" if is_import else " (Copia)"),
            'unita_misura': src_rec[4],
            'perc_spese_generali': src_rec[5],
            'perc_sicurezza': src_rec[6],
            'perc_utili': src_rec[7],
            'prezzo_finale': src_rec[8]
        }

        try:
            new_np_id = self.db.insert("nuovi_prezzi", data)
            items = self.db.fetch_all("voci_costo", "WHERE np_id=?", (source_id,))
            for item in items:
                new_item_data = {
                    'np_id': new_np_id,
                    'ordine': item[2],
                    'categoria': item[3],
                    'descrizione': item[4],
                    'um': item[5],
                    'quantita': item[6],
                    'prezzo_unitario': item[7]
                }
                self.db.insert("voci_costo", new_item_data)
            
            if self.notebook.index("current") == 1:
                self.tab_np.refresh_data("WHERE progetto_id=?", (self.current_project_id,))
            messagebox.showinfo("Successo", "NP copiato con successo!")
        except Exception as e:
            messagebox.showerror("Errore Copia", str(e))

    # --- EVENT HANDLERS ---
    def on_tab_change(self, event):
        idx = self.notebook.index("current")
        if idx == 1 and self.current_project_id:
            self.tab_np.refresh_data("WHERE progetto_id=?", (self.current_project_id,))
        elif idx == 4:
            current_np = self.current_np_id if hasattr(self, 'current_np_id') else None
            self.tab_print.set_current_np(current_np)
        elif idx == 5: 
            self.tab_converter.refresh_file_list()

    def on_project_select(self, project_id):
        self.current_project_id = project_id
        proj = self.db.fetch_one("progetti", "WHERE id=?", (project_id,))
        if proj:
            self.lbl_project_title.config(text=f"Progetto Attivo: {proj[2]} (Cod: {proj[1]})")
        
        self.tab_np.refresh_data("WHERE progetto_id=?", (project_id,))
        self.current_np_id = None
        self.clear_details_view()

    def goto_tab_np(self, project_id):
        self.on_project_select(project_id)
        self.notebook.select(self.frame_tab2)

    def on_np_select(self, np_id):
        self.current_np_id = np_id
        self.refresh_details_tree()

    def goto_tab_details(self, np_id):
        self.on_np_select(np_id)
        self.notebook.select(self.frame_det_np)

    # --- LOGICA TAB 3 ---
    def setup_details_tab(self):
        self.lbl_np_details = ttk.Label(self.frame_det_np, text="Nessun NP Selezionato", font=("Arial", 12, "bold"), foreground="#0055aa", padding=10)
        self.lbl_np_details.pack(side="top", fill="x")

        # Input Frame
        self.det_input_frame = ttk.LabelFrame(self.frame_det_np, text="Voce di Costo", padding=10)
        self.det_input_frame.pack(side="top", fill="x", padx=10, pady=5)

        # --- RIGA 1: Intestazioni ---
        ttk.Label(self.det_input_frame, text="N. Ordine", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky="w", padx=5)
        ttk.Label(self.det_input_frame, text="Categoria", font=("Arial", 9, "bold")).grid(row=0, column=1, sticky="w", padx=5)
        ttk.Label(self.det_input_frame, text="Descrizione", font=("Arial", 9, "bold")).grid(row=0, column=2, sticky="w", padx=5)

        # --- RIGA 2: Input Parte Alta ---
        # Ordine
        self.entry_order = ttk.Entry(self.det_input_frame, width=8)
        self.entry_order.grid(row=1, column=0, sticky="nw", padx=5, pady=(0, 10))

        # Categoria (MODIFICATA: Solo Manodopera, Prodotti, Attrezzature)
        self.combo_cat = ttk.Combobox(self.det_input_frame, values=["Manodopera", "Prodotti", "Attrezzature"], state="readonly", width=18)
        self.combo_cat.grid(row=1, column=1, sticky="nw", padx=5, pady=(0, 10))
        self.combo_cat.current(0)

        # Descrizione (Frame Contenitore per Text + Scrollbar)
        desc_frame = ttk.Frame(self.det_input_frame)
        desc_frame.grid(row=1, column=2, sticky="ew", padx=5, pady=(0, 10))
        
        self.txt_desc = tk.Text(desc_frame, height=2, width=30, font=("Arial", 9))
        self.txt_desc.pack(side="left", fill="both", expand=True)
        
        sb_desc = ttk.Scrollbar(desc_frame, orient="vertical", command=self.txt_desc.yview)
        sb_desc.pack(side="right", fill="y")
        self.txt_desc.config(yscrollcommand=sb_desc.set)
        
        self.det_input_frame.columnconfigure(2, weight=1)

        # --- RIGA 3: Intestazioni Parte Bassa ---
        ttk.Label(self.det_input_frame, text="UM", font=("Arial", 9, "bold")).grid(row=2, column=0, sticky="w", padx=5)
        ttk.Label(self.det_input_frame, text="Quantità", font=("Arial", 9, "bold")).grid(row=2, column=1, sticky="w", padx=5)
        ttk.Label(self.det_input_frame, text="Prezzo Unitario", font=("Arial", 9, "bold")).grid(row=2, column=2, sticky="w", padx=5)

        # --- RIGA 4: Input Parte Bassa ---
        self.combo_um = ttk.Combobox(self.det_input_frame, width=12, postcommand=self.populate_um_combo) 
        self.combo_um.grid(row=3, column=0, sticky="w", padx=5, pady=(0, 5))

        self.entry_q = ttk.Entry(self.det_input_frame, width=12)
        self.entry_q.grid(row=3, column=1, sticky="w", padx=5, pady=(0, 5))

        self.entry_pu = ttk.Entry(self.det_input_frame, width=12)
        self.entry_pu.grid(row=3, column=2, sticky="w", padx=5, pady=(0, 5))

        # Content Frame
        content_frame = ttk.Frame(self.frame_det_np)
        content_frame.pack(side="top", fill="both", expand=True, padx=10, pady=5)

        cols = ("ordine", "cat", "desc", "um", "q", "pu", "totale")
        self.tree_det = ttk.Treeview(content_frame, columns=cols, show="headings")
        headers = ["N.", "Categoria", "Descrizione", "UM", "Q", "PU", "Totale"]
        widths = [40, 100, 300, 50, 80, 80, 100]
        
        # Setup Ordinamento Tab 3
        for c, h, w in zip(cols, headers, widths):
            align = "w"
            if c in ("ordine", "q", "pu", "totale"): align = "e"
            self.tree_det.heading(c, text=h, command=lambda _c=c: self.sort_det_column(_c, False))
            self.tree_det.column(c, width=w, anchor=align)
        
        self.tree_det.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(content_frame, orient="vertical", command=self.tree_det.yview)
        sb.pack(side="left", fill="y")
        self.tree_det.configure(yscroll=sb.set)

        btn_frame = ttk.Frame(content_frame)
        btn_frame.pack(side="right", fill="y", padx=5)
        ttk.Button(btn_frame, text="Aggiungi", command=self.add_cost_item).pack(fill="x", pady=2)
        ttk.Button(btn_frame, text="Aggiorna", command=self.update_cost_item).pack(fill="x", pady=2)
        ttk.Button(btn_frame, text="Elimina", command=self.delete_cost_item).pack(fill="x", pady=2)
        ttk.Button(btn_frame, text="Pulisci", command=self.clear_details_inputs).pack(fill="x", pady=10)
        
        # --- PULSANTE NUOVO: CERCA DA STORICO ---
        ttk.Separator(btn_frame, orient="horizontal").pack(fill="x", pady=5)
        ttk.Button(btn_frame, text="Cerca da Storico", command=self.open_history_dialog).pack(fill="x", pady=2)

        self.tree_det.bind("<<TreeviewSelect>>", self.on_det_select)

        # RIEPILOGO ECONOMICO
        self.frame_summary = ttk.LabelFrame(self.frame_det_np, text="Riepilogo Economico", padding=15)
        self.frame_summary.pack(side="bottom", fill="x", padx=10, pady=10)
        
        self.frame_summary.columnconfigure(0, weight=0)
        self.frame_summary.columnconfigure(1, weight=0) 
        self.frame_summary.columnconfigure(2, weight=0)
        self.frame_summary.columnconfigure(3, weight=1)

        ttk.Label(self.frame_summary, text="(A) Totale Voci Elementari:", font=("Arial", 10)).grid(row=0, column=0, columnspan=3, sticky="e", padx=5, pady=2)
        self.lbl_sum_a = ttk.Label(self.frame_summary, text="€ 0.00", font=("Arial", 10))
        self.lbl_sum_a.grid(row=0, column=3, sticky="e", padx=5, pady=2)

        ttk.Label(self.frame_summary, text="Spese Generali:", font=("Arial", 10)).grid(row=1, column=0, sticky="e", padx=5, pady=2)
        self.entry_perc_spese = ttk.Entry(self.frame_summary, width=6, justify="right")
        self.entry_perc_spese.grid(row=1, column=1, sticky="w", pady=2)
        self.entry_perc_spese.bind("<Return>", lambda e: self.recalculate_totals())
        self.entry_perc_spese.bind("<FocusOut>", lambda e: self.recalculate_totals())
        ttk.Label(self.frame_summary, text="%").grid(row=1, column=2, sticky="w")
        self.lbl_sum_b = ttk.Label(self.frame_summary, text="€ 0.00 (B)", font=("Arial", 10))
        self.lbl_sum_b.grid(row=1, column=3, sticky="e", padx=5, pady=2)

        ttk.Label(self.frame_summary, text="↳ di cui Sicurezza (inclusa):", font=("Arial", 9), foreground="#555").grid(row=2, column=0, sticky="e", padx=5, pady=0)
        self.entry_perc_sicurezza = ttk.Entry(self.frame_summary, width=6, justify="right")
        self.entry_perc_sicurezza.grid(row=2, column=1, sticky="w", pady=0)
        self.entry_perc_sicurezza.bind("<Return>", lambda e: self.recalculate_totals())
        self.entry_perc_sicurezza.bind("<FocusOut>", lambda e: self.recalculate_totals())
        ttk.Label(self.frame_summary, text="%", foreground="#555").grid(row=2, column=2, sticky="w")
        self.lbl_val_sicurezza = ttk.Label(self.frame_summary, text="€ 0.00", font=("Arial", 9), foreground="#555")
        self.lbl_val_sicurezza.grid(row=2, column=3, sticky="e", padx=5, pady=0)

        ttk.Label(self.frame_summary, text="Utili d'Impresa:", font=("Arial", 10)).grid(row=3, column=0, sticky="e", padx=5, pady=2)
        self.entry_perc_utili = ttk.Entry(self.frame_summary, width=6, justify="right")
        self.entry_perc_utili.grid(row=3, column=1, sticky="w", pady=2)
        self.entry_perc_utili.bind("<Return>", lambda e: self.recalculate_totals())
        self.entry_perc_utili.bind("<FocusOut>", lambda e: self.recalculate_totals())
        ttk.Label(self.frame_summary, text="%").grid(row=3, column=2, sticky="w")
        self.lbl_sum_c = ttk.Label(self.frame_summary, text="€ 0.00 (C)", font=("Arial", 10))
        self.lbl_sum_c.grid(row=3, column=3, sticky="e", padx=5, pady=2)

        ttk.Separator(self.frame_summary, orient="horizontal").grid(row=4, column=0, columnspan=4, sticky="ew", pady=10)

        ttk.Label(self.frame_summary, text="TOTALE COMPLESSIVO NP:", font=("Arial", 12, "bold")).grid(row=5, column=0, columnspan=3, sticky="e", padx=5)
        self.lbl_total_final = ttk.Label(self.frame_summary, text="€ 0.00", font=("Arial", 13, "bold"), foreground="#0055aa")
        self.lbl_total_final.grid(row=5, column=3, sticky="e", padx=5)

        self.lbl_incidences = ttk.Label(self.frame_summary, text="", font=("Arial", 8), foreground="#777")
        self.lbl_incidences.grid(row=6, column=0, columnspan=4, sticky="w", pady=(5,0))

    def sort_det_column(self, col, reverse):
        # Ordinamento Tab 3 con rilevamento tipo dato
        l = [(self.tree_det.set(k, col), k) for k in self.tree_det.get_children('')]
        
        def det_sort_key(t):
            val = t[0]
            if not val: return ""
            try:
                # Gestione numeri formattati es. "1.000,00" o "100"
                clean = str(val).replace('€', '').strip()
                if '.' in clean and ',' in clean: clean = clean.replace('.', '').replace(',', '.')
                elif ',' in clean: clean = clean.replace(',', '.')
                return float(clean)
            except ValueError:
                return str(val).lower()

        l.sort(key=det_sort_key, reverse=reverse)
        for index, (val, k) in enumerate(l):
            self.tree_det.move(k, '', index)
        self.tree_det.heading(col, command=lambda: self.sort_det_column(col, not reverse))

    def populate_um_combo(self):
        ums = self.db.fetch_all("unita_misura")
        values = [u[1] for u in ums]
        self.combo_um['values'] = values

    def clear_details_inputs(self):
        self.entry_order.delete(0, tk.END)
        # Svuota il widget Text
        self.txt_desc.delete("1.0", tk.END)
        self.entry_q.delete(0, tk.END)
        self.entry_pu.delete(0, tk.END)
        self.combo_cat.current(0)
        self.combo_um.set('')
        if self.current_np_id:
            max_ord = self.db.get_max_order(self.current_np_id)
            self.entry_order.insert(0, str(max_ord + 1))

    def clear_details_view(self):
        for row in self.tree_det.get_children(): self.tree_det.delete(row)
        self.lbl_sum_a.config(text="€ 0.00")
        self.lbl_sum_b.config(text="€ 0.00")
        self.lbl_sum_c.config(text="€ 0.00")
        self.lbl_val_sicurezza.config(text="€ 0.00")
        self.lbl_total_final.config(text="€ 0.00")
        self.lbl_incidences.config(text="")
        self.lbl_np_details.config(text="Nessun NP Selezionato")
        self.clear_details_inputs()

    def refresh_details_tree(self):
        for row in self.tree_det.get_children(): self.tree_det.delete(row)
        if not hasattr(self, 'current_np_id') or not self.current_np_id: return
        
        np_rec = self.db.fetch_one("nuovi_prezzi", "WHERE id=?", (self.current_np_id,))
        if np_rec:
            self.lbl_np_details.config(text=f"NP Attivo: {np_rec[2]} - {np_rec[3]}")
            val_spese = np_rec[5] if np_rec[5] is not None else 17.0
            val_sicurezza = np_rec[6] if np_rec[6] is not None else 5.0
            val_utili = np_rec[7] if np_rec[7] is not None else 10.0
            self.entry_perc_spese.delete(0, tk.END); self.entry_perc_spese.insert(0, str(val_spese))
            self.entry_perc_sicurezza.delete(0, tk.END); self.entry_perc_sicurezza.insert(0, str(val_sicurezza))
            self.entry_perc_utili.delete(0, tk.END); self.entry_perc_utili.insert(0, str(val_utili))

        items = self.db.fetch_all("voci_costo", "WHERE np_id=? ORDER BY ordine ASC", (self.current_np_id,))
        self.total_a = 0.0
        self.total_manodopera_base = 0.0
        max_ord = 0

        for item in items:
            ord_val, cat, desc, um, q, pu = item[2], item[3], item[4], item[5], item[6], item[7]
            if ord_val > max_ord: max_ord = ord_val
            tot = round(q * pu, 2)
            self.total_a += tot
            if "Manodopera" in cat: self.total_manodopera_base += tot
            q_fmt = f"{q:.3f}".replace('.', ',')
            pu_fmt = f"{pu:.3f}".replace('.', ',')
            tot_fmt = f"{tot:.2f}".replace('.', ',')
            self.tree_det.insert("", "end", iid=item[0], values=(ord_val, cat, desc, um, q_fmt, pu_fmt, tot_fmt))
            
        self.entry_order.delete(0, tk.END)
        self.entry_order.insert(0, str(max_ord + 1))
        self.recalculate_totals()

    def recalculate_totals(self):
        if not self.current_np_id: return
        try:
            p_spese = safe_float_convert(self.entry_perc_spese.get())
            p_sicurezza = safe_float_convert(self.entry_perc_sicurezza.get())
            p_utili = safe_float_convert(self.entry_perc_utili.get())
        except: return

        val_a = self.total_a
        val_b = round(val_a * (p_spese / 100.0), 2)
        val_c = round((val_a + val_b) * (p_utili / 100.0), 2)
        val_total = val_a + val_b + val_c
        val_sicurezza = round(val_b * (p_sicurezza / 100.0), 2)
        
        perc_sic_tot = (val_sicurezza / val_total * 100) if val_total > 0 else 0
        perc_man_tot = (self.total_manodopera_base / val_total * 100) if val_total > 0 else 0

        self.lbl_sum_a.config(text=format_currency(val_a))
        self.lbl_sum_b.config(text=format_currency(val_b))
        self.lbl_val_sicurezza.config(text=f"({format_currency(val_sicurezza)})")
        self.lbl_sum_c.config(text=format_currency(val_c))
        self.lbl_total_final.config(text=format_currency(val_total))
        self.lbl_incidences.config(text=f"INFO: Incidenza Manodopera sul totale: {perc_man_tot:.2f}% | Incidenza Sicurezza sul totale: {perc_sic_tot:.2f}%")

        data = {'perc_spese_generali': p_spese, 'perc_sicurezza': p_sicurezza, 'perc_utili': p_utili, 'prezzo_finale': val_total}
        self.db.update("nuovi_prezzi", self.current_np_id, data)

    def on_det_select(self, event):
        selected = self.tree_det.selection()
        if not selected: return
        vals = self.tree_det.item(selected[0])['values']
        self.entry_order.delete(0, tk.END); self.entry_order.insert(0, vals[0])
        self.combo_cat.set(vals[1])
        
        # Gestione Text widget
        self.txt_desc.delete("1.0", tk.END)
        self.txt_desc.insert("1.0", vals[2])
        
        self.combo_um.set(vals[3])
        q_val = safe_float_convert(vals[4])
        pu_val = safe_float_convert(vals[5])
        self.entry_q.delete(0, tk.END); 
        self.entry_q.insert(0, f"{q_val:.3f}".replace('.', ','))
        self.entry_pu.delete(0, tk.END); 
        self.entry_pu.insert(0, f"{pu_val:.3f}".replace('.', ','))

    # --- FUNZIONI PER POPOLAMENTO DA STORICO ---
    def open_history_dialog(self):
        if not self.current_project_id:
            messagebox.showwarning("Attenzione", "Nessun progetto selezionato.")
            return
        HistoryDialog(self, self.db, self.current_project_id, self.import_from_history_wrapper)

    def import_from_history_wrapper(self, cat, desc, um, pu):
        # Popola i campi della UI
        self.combo_cat.set(cat)
        
        self.txt_desc.delete("1.0", tk.END)
        self.txt_desc.insert("1.0", desc)
        
        self.combo_um.set(um)
        
        self.entry_pu.delete(0, tk.END)
        self.entry_pu.insert(0, f"{pu:.3f}".replace('.', ','))
        
        # Pulisci quantità perché è specifica del nuovo inserimento
        self.entry_q.delete(0, tk.END)

    def add_cost_item(self):
        if not hasattr(self, 'current_np_id') or not self.current_np_id:
            messagebox.showwarning("Attenzione", "Seleziona NP")
            return
        try:
            q = safe_float_convert(self.entry_q.get())
            pu = safe_float_convert(self.entry_pu.get())
            ord_val = int(float(self.entry_order.get()))
        except ValueError as e:
            messagebox.showerror("Errore", str(e)); return
        
        # Recupero testo dal widget Text
        desc_val = self.txt_desc.get("1.0", "end-1c")
        
        data = {'np_id': self.current_np_id, 'ordine': ord_val, 'categoria': self.combo_cat.get(), 'descrizione': desc_val, 'um': self.combo_um.get(), 'quantita': q, 'prezzo_unitario': pu}
        self.db.insert("voci_costo", data)
        self.refresh_details_tree()

    def update_cost_item(self):
        selected = self.tree_det.selection()
        if not selected: return
        try:
            q = safe_float_convert(self.entry_q.get())
            pu = safe_float_convert(self.entry_pu.get())
            ord_val = int(float(self.entry_order.get()))
            
            # Recupero testo dal widget Text
            desc_val = self.txt_desc.get("1.0", "end-1c")
            
            data = {'ordine': ord_val, 'categoria': self.combo_cat.get(), 'descrizione': desc_val, 'um': self.combo_um.get(), 'quantita': q, 'prezzo_unitario': pu}
            self.db.update("voci_costo", selected[0], data)
            self.refresh_details_tree()
        except ValueError as e:
            messagebox.showerror("Errore", str(e))

    def delete_cost_item(self):
        selected = self.tree_det.selection()
        if not selected: return
        self.db.delete("voci_costo", selected[0])
        self.refresh_details_tree()

if __name__ == "__main__":
    app = NPApp()
    app.mainloop()